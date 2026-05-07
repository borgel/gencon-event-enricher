"""Parse a GenCon events xlsx dump into SessionRecord instances."""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .types import SessionRecord


def _parse_dt(value: Any) -> datetime | None:
    """GenCon writes start/end as either Excel serial floats or 'MM/DD/YYYY HH:MM AM' strings."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date — openpyxl normally converts these, but be defensive.
        from openpyxl.utils.datetime import from_excel
        return from_excel(value)
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%m/%d/%Y %I:%M %p")
    raise ValueError(f"unparseable datetime: {value!r}")


def _to_int(v: Any) -> int | None:
    if v in (None, ""):
        return None
    return int(v)


def _to_float(v: Any) -> float | None:
    if v in (None, ""):
        return None
    return float(v)


def _to_bool_yesno(v: Any) -> bool:
    return str(v).strip().lower() == "yes"


def parse_gencon_xlsx(path: Path) -> list[SessionRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(it)]
    idx = {h: i for i, h in enumerate(headers)}

    def col(row: tuple, name: str) -> Any:
        return row[idx[name]]

    out: list[SessionRecord] = []
    for row in it:
        if row is None or all(c is None or c == "" for c in row):
            continue
        out.append(SessionRecord(
            gencon_id=str(col(row, "Game ID")),
            group_label=str(col(row, "Group") or ""),
            title=str(col(row, "Title") or "").strip(),
            short_description=str(col(row, "Short Description") or ""),
            long_description=str(col(row, "Long Description") or ""),
            event_type=str(col(row, "Event Type") or "").strip(),
            game_system=str(col(row, "Game System") or "").strip(),
            rules_edition=str(col(row, "Rules Edition") or ""),
            min_players=_to_int(col(row, "Minimum Players")),
            max_players=_to_int(col(row, "Maximum Players")),
            age_required=str(col(row, "Age Required") or ""),
            experience_required=str(col(row, "Experience Required") or ""),
            materials_required=str(col(row, "Materials Required") or ""),
            materials_required_details=str(col(row, "Materials Required Details") or ""),
            start=_parse_dt(col(row, "Start Date & Time")),
            duration_minutes=_minutes_from_hours(col(row, "Duration")),
            end=_parse_dt(col(row, "End Date & Time")),
            gm_names=str(col(row, "GM Names") or ""),
            website=str(col(row, "Website") or ""),
            email=str(col(row, "Email") or ""),
            tournament=_to_bool_yesno(col(row, "Tournament?")),
            round_number=_to_int(col(row, "Round Number")),
            total_rounds=_to_int(col(row, "Total Rounds")),
            minimum_play_time=_to_int(col(row, "Minimum Play Time")),
            attendee_registration=str(col(row, "Attendee Registration?") or ""),
            cost=_to_float(col(row, "Cost $")),
            location=str(col(row, "Location") or ""),
            room=str(col(row, "Room Name") or ""),
            table=str(col(row, "Table Number") or ""),
            special_category=str(col(row, "Special Category") or ""),
            tickets_available=_to_int(col(row, "Tickets Available")),
            last_modified=_parse_dt(col(row, "Last Modified")),
        ))
    return out


def _minutes_from_hours(v: Any) -> int | None:
    """The xlsx 'Duration' column is hours (often a float). Convert to minutes."""
    if v in (None, ""):
        return None
    return int(round(float(v) * 60))
